# -*- coding: utf-8 -*-
"""诊断编码字典构建脚本（scripts/build_diagnosis_dict.py，2026-08-21 阶段2）。

输入（官方渠道原始 Excel，存 D:/文件/项目数据/MediScribe/编码字典/raw/）：
  - hubei_icd10_v2.xlsx  湖北省卫健委 2.0 版疾病+手术完整库（含医保对应码列）
    来源 http://wjw.hubei.gov.cn/bmdt/ywdt/yzyg/202102/t20210220_3354180.shtml
  - yn_*793463.xlsx      中医疾病名代码（GB/T 15657-2021，A 码）
  - yn_*433697.xlsx      中医证候名代码（GB/T 15657-2021，B 码）
    来源 云南省医保局医保版中医病证分类与代码更新材料（医保办函〔2021〕19号配套）

输出：backend/seed_data/diagnosis_codes.csv.gz（进 git 的代码资产），列：
  code, name, code_type(ICD10/ICD9CM3/TCD_DIS/TCD_SYN), pinyin_initial, aliases

清洗规则：
  - ICD10/手术：以「医保对应码/医保对应名称」为准（我们要医保 2.0 口径），
    医保码为空的行（地方增补码）跳过；同码去重取首行
  - 中医：尾缀带 '.' 的是类目标题（不可选择），跳过；
    名称含「（可选词：X；Y）」拆出主名 + 别名（检索时都可命中）
  - pinyin_initial：主名的拼音首字母串（pypinyin），检索联想用

跑法（本地一次性）：backend/venv/Scripts/python.exe scripts/build_diagnosis_dict.py <raw_dir>
产物幂等——同输入必同输出，重跑安全。
"""
import csv
import gzip
import io
import re
import sys
from pathlib import Path

import openpyxl
from pypinyin import Style, lazy_pinyin

OUT_PATH = Path(__file__).resolve().parents[1] / "seed_data" / "diagnosis_codes.csv.gz"

_ALIAS_RE = re.compile(r"[（(]可选词[：:]\s*(.+?)[)）]")


def _clean_name(raw: str) -> tuple[str, str]:
    """拆主名与可选词别名：'感冒（可选词：伤风感冒；时感）' → ('感冒', '伤风感冒；时感')。"""
    raw = (raw or "").strip()
    m = _ALIAS_RE.search(raw)
    aliases = ""
    if m:
        aliases = m.group(1).strip().replace(";", "；")
        raw = _ALIAS_RE.sub("", raw).strip()
    return raw, aliases


def _pinyin_initial(name: str) -> str:
    """主名 → 拼音首字母串（小写；数字/字母原样保留）。

    2026-08-22 终验修正：lazy_pinyin(errors='ignore') 会丢掉非汉字——
    "2型糖尿病"变成 "xtnb"，医生输 "2xtnb" 搜不到。逐字符处理：
    汉字取拼音首字母，ASCII 数字/字母原样保留。
    """
    out = []
    for ch in name:
        if ch.isascii():
            if ch.isalnum():
                out.append(ch.lower())
        else:
            letters = lazy_pinyin(ch, style=Style.FIRST_LETTER, errors="ignore")
            if letters and letters[0]:
                out.append(letters[0][0].lower())
    return "".join(out)


def _load_hubei(path: Path) -> list[tuple[str, str, str]]:
    """湖北完整库 → [(code, name, code_type)]，以医保对应码为准。"""
    wb = openpyxl.load_workbook(path, read_only=True)
    out: list[tuple[str, str, str]] = []
    sheet_types = {
        "湖北2.0版疾病完整库（含医保映射）": "ICD10",
        "湖北2.0版手术完整库（含医保映射）": "ICD9CM3",
    }
    for sheet, ctype in sheet_types.items():
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, values_only=True):
            yb_code = str(row[3] or "").strip()
            yb_name = str(row[4] or "").strip()
            if not yb_code or not yb_name:
                continue  # 医保未收录的地方增补码——跳过（我们要医保口径）
            out.append((yb_code, yb_name, ctype))
    return out


def _load_tcm(path: Path, ctype: str) -> list[tuple[str, str, str, str]]:
    """云南中医表 → [(code, name, ctype, aliases)]；尾点=类目跳过。"""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    out: list[tuple[str, str, str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = str(row[1] or "").strip()
        raw_name = str(row[2] or "").strip()
        if not code or not raw_name:
            continue
        if code.endswith("."):
            continue  # 类目标题行，不是可选择的诊断
        name, aliases = _clean_name(raw_name)
        if name:
            out.append((code, name, ctype, aliases))
    return out


def main(raw_dir: str) -> None:
    raw = Path(raw_dir)
    hubei = next(raw.glob("hubei_icd10_v2*.xlsx"))
    tcm_dis = next(raw.glob("*793463*.xlsx"))
    tcm_syn = next(raw.glob("*433697*.xlsx"))

    rows: list[tuple[str, str, str, str, str]] = []
    seen: set[tuple[str, str]] = set()

    for code, name, ctype in _load_hubei(hubei):
        key = (ctype, code)
        if key in seen:
            continue
        seen.add(key)
        rows.append((code, name, ctype, _pinyin_initial(name), ""))

    for path, ctype in ((tcm_dis, "TCD_DIS"), (tcm_syn, "TCD_SYN")):
        for code, name, ct, aliases in _load_tcm(path, ctype):
            key = (ct, code)
            if key in seen:
                continue
            seen.add(key)
            rows.append((code, name, ct, _pinyin_initial(name), aliases))

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["code", "name", "code_type", "pinyin_initial", "aliases"])
    writer.writerows(rows)
    with gzip.open(OUT_PATH, "wt", encoding="utf-8", newline="") as f:
        f.write(buf.getvalue())

    by_type: dict[str, int] = {}
    for r in rows:
        by_type[r[2]] = by_type.get(r[2], 0) + 1
    print(f"写出 {OUT_PATH}：共 {len(rows)} 条 -> {by_type}")


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else ".")
